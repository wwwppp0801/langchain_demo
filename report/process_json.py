import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def _convert_json_to_excel(input_file, output_file):
    df = pd.read_json(input_file)
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle

    #red_fill = PatternFill(bgColor="FFC7CE")
    #green_fill = PatternFill(bgColor="C6EFCE")


    validate_result_col=None
    # 遍历第一行的所有单元格
    for cell in ws[1]:
        # 如果单元格的值为“validate_result”，则获取该单元格的列索引
        if cell.value == 'validate_result':
            validate_result_col = cell.column

    # 设置填充颜色
    trueFill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    falseFill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    #trueFill = PatternFill(bgColor="FFC7CE")
    #falseFill = PatternFill(bgColor="C6EFCE")

    # 遍历行
    if validate_result_col:
        for row in ws.rows:
            # 获取validate_result列的值
            validate_result = row[validate_result_col - 1].value

            # 如果validate_result为TRUE，则将整行的背景设置为绿色
            if str(validate_result).lower() == 'true':
                for cell in row:
                    cell.fill = trueFill
            # 如果validate_result为FALSE，则将整行的背景设置为红色
            elif str(validate_result).lower() == 'false':
                for cell in row:
                    cell.fill = falseFill

    # 遍历第一行单元格，找到raw_result所在的列号
    for cell in ws[1]:
        if cell.value == "raw_result":
            ws.column_dimensions[cell.column_letter].width = 80
        if cell.value == "final_result":
            ws.column_dimensions[cell.column_letter].width = 30
        if cell.value == "validator":
            ws.column_dimensions[cell.column_letter].width = 20



    wb.save(output_file)

import traceback

def convert_json_to_excel(input_file,output_file):
    try:
        return _convert_json_to_excel(input_file,output_file)
    except Exception as e:
        traceback.print_exc()
        return None

def traverse_directory(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.json'):
                input_file = os.path.join(root, file)
                output_file = os.path.splitext(input_file)[0] + '.xlsx'
                convert_json_to_excel(input_file, output_file)


if __name__=="__main__":
    if len(sys.argv)==2:
        path=sys.argv[1]
        if os.path.isdir(path):
            traverse_directory(path)
        elif os.path.isfile(path):
            convert_json_to_excel(path,path+".xlsx")
        else:
            print("error:"+path+" is not file or dir")
    else:
        traverse_directory('.')



